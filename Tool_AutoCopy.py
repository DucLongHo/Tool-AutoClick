import nodriver as uc
import asyncio
import pandas as pd
import os
from openpyxl import load_workbook
import pyautogui

# --- HẰNG SỐ CẤU HÌNH ---
TIMECOUNT = 5
LOGIN_URL = "https://secure.vantagemarkets.com/login"
EXCEL_FILE = "accounts.xlsx"
MAX_CONCURRENT = 6 # Số cửa sổ mở cùng lúc

screen_width, screen_height = pyautogui.size()
COLS = 3
ROWS = 2

WIDTH = screen_width // COLS   
HEIGHT = screen_height // ROWS 

async def login_account(email, password, index):
    quadrant = index % 6
    x = (quadrant % COLS) * WIDTH
    y = (quadrant // COLS) * HEIGHT
    
    args = [
        f"--window-size={WIDTH},{HEIGHT}",
        f"--window-position={x},{y}",
        "--no-first-run",
    ]

    browser = await uc.start(browser_args=args, headless=False)
    status = ""

    try:
        page = await browser.get(LOGIN_URL)        
        await asyncio.sleep(TIMECOUNT)

        # 1. Nhập Email
        try:
            email_field = await page.wait_for('input[data-testid="userName_login"]', timeout=TIMECOUNT)
            await email_field.send_keys(email)
        except Exception:
            return email, "ERROR: Không tìm thấy ô Email."

        # 2. Nhập Password
        try:
            pass_field = await page.select('input[type="password"]') 
            await pass_field.send_keys(password)
        except Exception:
            return email, "ERROR: Không tìm thấy ô Password."


        # 3. Click Đăng nhập
        try:
            login_btn = await page.wait_for('button[data-testid="login"]', timeout=TIMECOUNT)
            await login_btn.click()
        except Exception:
            return email, "ERROR: Không tìm thấy nút Login."
        
        # 4. Kiểm tra URL xem đã qua được trang login chưa
        await asyncio.sleep(TIMECOUNT)
        if "login" not in page.url.lower():
            status = "Success"
        else:
            return email, "Failed: Sai mật khẩu hoặc bị chặn"
            
    except Exception as e:
        status = f"ERROR: {str(e)[:20]}"
    finally:
        browser.stop()
        return email, status

def update_all_excel_status(results):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # Tìm index của cột Email và Status
        header = [cell.value for cell in ws[1]]
        try:
            email_col = header.index("Email") + 1
            status_col = header.index("Status") + 1
        except ValueError:
            print("[!] Lỗi: Không tìm thấy cột 'Email' hoặc 'Status' trong Excel!")
            return

        # Chuyển list results thành dictionary {email: status} để tra cứu
        result_dict = dict(results)

        # Ghi data vào file
        for row in range(2, ws.max_row + 1):
            cell_email = str(ws.cell(row=row, column=email_col).value).strip()
            if cell_email in result_dict:
                ws.cell(row=row, column=status_col).value = result_dict[cell_email]
        
        wb.save(EXCEL_FILE)
        print(f"[*] Đã lưu thành công toàn bộ trạng thái vào {EXCEL_FILE}")
    except Exception as e:
        print(f"[!] Lỗi khi ghi Excel: {e}")

async def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"ERROR: Không tìm thấy file {EXCEL_FILE}")
        return

    df = pd.read_excel(EXCEL_FILE)
    print(f"[*] Tìm thấy {len(df)} tài khoản. Đang khởi chạy tối đa {MAX_CONCURRENT} luồng...")
    print("-" * 30)

    # Dùng Semaphore để giới hạn số luồng (mở cùng lúc 2 cửa sổ, đóng cái nào mở bù cái đó)
    semaphore = asyncio.Semaphore(MAX_CONCURRENT)
    tasks = []

    async def safe_login(email, password, index):
        async with semaphore:
            return await login_account(email, password, index)

    for index, row in df.iterrows():
        email = str(row['Email']).strip()
        password = str(row['Password']).strip()
        tasks.append(safe_login(email, password, index))

    results = await asyncio.gather(*tasks)

    # In kết quả ra màn hình
    for email, res in results:
        print(f"[>] {email}: {res}")
    print("=" * 30)

    # GỌI HÀM LƯU EXCEL ĐÃ VIẾT (Bản cũ bạn quên gọi hàm này)
    if results:
        print("Đang ghi vào file Excel...")
        update_all_excel_status(results)

    print("[FINISH] Hoàn thành toàn bộ quy trình.")

if __name__ == "__main__":
    uc.loop().run_until_complete(main())